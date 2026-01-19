import os
import re
import pandas as pd
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
import requests

# ---------------- HELPER ----------------

def working_days_between(start, end):
    days = 0
    current = start
    while current <= end:
        if current.weekday() < 5:
            days += 1
        current += timedelta(days=1)
    return days

# ---------------- FETCH IPO DATA ----------------

def get_ipos():
    print("-- Starting IPO data extraction")
    today = datetime.today().date()

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    print("-- Launching Chrome WebDriver in headless mode")
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 30)

    print("-- Navigating to IPO GMP report page")
    driver.get("https://www.investorgain.com/report/live-ipo-gmp/331/all/")

    print("-- Waiting for IPO table to load")
    table = wait.until(EC.presence_of_element_located((By.ID, "report_table")))
    rows = table.find_elements(By.TAG_NAME, "tr")

    ipo_data = []
    print("-- Extracting IPO rows")

    for row in rows[1:]:
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) > 8:
            name = cols[0].text.strip()
            gmp_text = cols[1].text.strip()
            sub = cols[3].text.strip()
            start = cols[7].text.strip()
            end = cols[8].text.strip()

            print(f"-- Processing IPO: {name}")

            match = re.search(r"\(([\d\.]+)%\)", gmp_text)
            gmp_value = float(match.group(1)) if match else 0

            try:
                def extract_date(text, today):
                    match = re.search(r"\d{1,2}-[A-Za-z]{3}", text)
                    if match:
                        return datetime.strptime(match.group(), "%d-%b").date().replace(year=today.year)
                    return None

                start_date = extract_date(start, today)
                end_date = extract_date(end, today)
                print(f"-- Extracted dates: Start={start_date}, End={end_date}")

            except Exception as e:
                print("-- Date extraction failed:", e)
                continue

            ipo_data.append((name, gmp_value, start_date, end_date, sub))

    driver.quit()
    print(f"-- IPO data extraction complete. Total IPOs found: {len(ipo_data)}")
    return ipo_data

# ---------------- TELEGRAM ----------------

def send_telegram_message(message):
    print("-- Preparing to send Telegram message")
    TELEGRAM_TOKEN = os.getenv("TG_BOT_TOKEN")
    TELEGRAM_CHAT_ID = os.getenv("TG_CHAT_ID")

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message}

    try:
        requests.post(url, data=payload, timeout=10)
        print("-- Telegram message sent successfully")
    except Exception as e:
        print("-- Telegram send failed:", e)

# ---------------- EXCEL UPDATE + VALIDATION ----------------

def update_excel(ipos):
    print("-- Starting Excel update process")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "TestData", "IPO_GMP.xlsx")
    print(f"-- Excel file path: {excel_file}")

    today = datetime.today().date()
    print(f"-- Today date: {today}")

    print("-- Reading Excel file into DataFrame")
    df = pd.read_excel(excel_file)

    print("-- Processing IPOs (ONLY 1 day before closing & GMP > 30)")

    for name, gmp, start, end, sub in ipos:

        if not end:
            print(f"-- Skipping IPO {name} due to missing end date")
            continue

        # ðŸ”¹ ONLY WHEN TODAY = ONE DAY BEFORE CLOSING
        if today == (end - timedelta(days=1)):
            print(f"-- IPO closing TOMORROW detected: {name}, GMP={gmp}")

            existing = df[df["IPO Name"] == name]

            if existing.empty:
                print(f"-- Adding new IPO {name}")
                df = pd.concat([df, pd.DataFrame([[name, gmp, start, end, sub, ""]],
                                                 columns=df.columns)], ignore_index=True)
            else:
                print(f"-- Updating GMP for IPO {name}")
                df.loc[df["IPO Name"] == name, "GMP"] = gmp

            # ðŸ”¹ RULE: GMP > 30 â†’ PROCEED
            try:
                if float(gmp) > 30:
                    print(f"-- GMP PASSED for {name}, GMP={gmp}")

                    df.loc[df["IPO Name"] == name, "Status"] = "Proceed"

                    message = (
                        f"ðŸš€ IPO PROCEED ALERT (Closing Tomorrow)\n\n"
                        f"Name: {name}\n"
                        f"GMP: {gmp}\n"
                        f"Subscription: {sub}\n"
                        f"Start Date: {start}\n"
                        f"End Date: {end}\n\n"
                        f"Status: PROCEED"
                    )

                    send_telegram_message(message)

                else:
                    print(f"-- GMP below 30 for {name}, GMP={gmp}")
                    df.loc[df["IPO Name"] == name, "Status"] = "Skip"

            except Exception as e:
                print(f"-- Error processing GMP for {name}: {e}")

        else:
            print(f"-- IPO {name} not in validation window (Today={today}, End={end}), skipping")

    # ðŸ”¹ Remove expired IPOs
    print("-- Cleaning up expired IPOs")
    df = df[pd.to_datetime(df["End Date"]).dt.date >= today]

    # ðŸ”¹ Save Excel
    print("-- Saving updated DataFrame back to Excel")
    df.to_excel(excel_file, index=False)

    # ðŸ”¹ Auto formatting
    print("-- Applying auto-formatting to Excel file")
    wb = load_workbook(excel_file)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(excel_file)
    print("-- Excel update process completed successfully")

# ---------------- RUN DAILY ----------------

print("-- Script execution started")
ipos = get_ipos()
update_excel(ipos)
print("-- Script execution finished")

